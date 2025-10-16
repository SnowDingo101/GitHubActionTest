#!/usr/bin/env python3
"""Updates an Excel file with timestamp and random data."""

from __future__ import annotations

import random
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    exit(1)


EXCEL_FILE = Path("data_report.xlsx")


def update_excel_file() -> str:
    """Add a new row with timestamp and random number to the Excel file."""
    
    # Create or load workbook
    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Report"
        # Add headers if new file
        ws.append(["Timestamp", "Random Number", "Status"])
    
    # Generate data
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    random_number = random.randint(1, 1000)
    status = "Active" if random_number > 500 else "Pending"
    
    # Add new row
    new_row = [timestamp, random_number, status]
    ws.append(new_row)
    
    # Save file
    wb.save(EXCEL_FILE)
    
    return f"Added row: {new_row}"


def main() -> None:
    result = update_excel_file()
    print(f"Excel file updated: {result}")
    print(f"File location: {EXCEL_FILE.absolute()}")


if __name__ == "__main__":
    main()