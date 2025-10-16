#!/usr/bin/env python3
"""Reads Excel file contents and sends them via email."""

from __future__ import annotations

import os
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    exit(1)

from email_sender import EmailConfig, send_email_smtp


EXCEL_FILE = Path("data_report.xlsx")


def read_excel_contents() -> str:
    """Read the Excel file and format contents for email."""
    
    if not EXCEL_FILE.exists():
        return "Excel file does not exist yet."
    
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Read all data
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows:
            return "Excel file is empty."
        
        # Format as text table
        content = f"Data Report Summary ({len(rows)} rows):\n\n"
        
        # Add header
        if rows:
            headers = rows[0]
            content += " | ".join(str(h) for h in headers) + "\n"
            content += "-" * (len(content.split('\n')[-2])) + "\n"
        
        # Add data rows (limit to last 10 for email brevity)
        data_rows = rows[1:] if len(rows) > 1 else []
        display_rows = data_rows[-10:] if len(data_rows) > 10 else data_rows
        
        for row in display_rows:
            content += " | ".join(str(cell) if cell is not None else "" for cell in row) + "\n"
        
        if len(data_rows) > 10:
            content += f"\n... (showing last 10 of {len(data_rows)} data rows)"
        
        return content
        
    except Exception as e:
        return f"Error reading Excel file: {e}"


def send_excel_report() -> None:
    """Send an email with Excel file contents."""
    
    # Read Excel data
    excel_content = read_excel_contents()
    
    # Email configuration (using same pattern as main.py)
    smtp_host = os.environ.get("SMTP_HOST", "")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    smtp_user = os.environ.get("SMTP_USERNAME", "")
    smtp_pass = os.environ.get("SMTP_PASSWORD", "")
    from_addr = os.environ.get("SMTP_FROM", smtp_user)
    use_ssl = os.environ.get("SMTP_USE_SSL", "").lower() in {"true", "1"}
    use_tls = os.environ.get("SMTP_USE_TLS", "true").lower() in {"true", "1"}
    dry_run = os.environ.get("DRY_RUN", "").lower() in {"true", "1"}
    
    # Email details
    to_email = "EMoroney@foresightgroupau.com"
    subject = f"Automated Data Report - {EXCEL_FILE.name}"
    body = f"""Hello,

This is an automated report from your GitHub Actions workflow.

{excel_content}

Best regards,
Automated Report System"""
    
    cfg = EmailConfig(
        host=smtp_host,
        port=smtp_port,
        username=smtp_user,
        password=smtp_pass,
        use_tls=use_tls,
        use_ssl=use_ssl,
        from_addr=from_addr,
        to_addrs=[to_email],
        subject=subject,
        body=body,
    )
    
    if dry_run or not smtp_host:
        print("[DRY RUN] Excel report email would be sent:")
        print(f"From: {cfg.from_addr}")
        print(f"To: {', '.join(cfg.to_addrs)}")
        print(f"Subject: {cfg.subject}")
        print("Body:")
        print(cfg.body)
        if not smtp_host:
            print("Note: Set SMTP_HOST to actually send the email.")
        return
    
    # Send the email
    try:
        send_email_smtp(cfg)
        print("Excel report email sent successfully.")
    except Exception as e:
        print(f"Failed to send Excel report email: {e}")


def main() -> None:
    send_excel_report()


if __name__ == "__main__":
    main()